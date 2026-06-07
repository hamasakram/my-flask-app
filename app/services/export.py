from io import BytesIO
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Image, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

LOGO_PATH = Path(__file__).resolve().parent.parent / "static" / "images" / "rn-colour-logo.png"
BRAND_HEADER_COLOR = "B21E22"


def _add_excel_logo(ws, start_row=1):
    if not LOGO_PATH.exists():
        return start_row

    logo = XLImage(str(LOGO_PATH))
    logo.width = 180
    logo.height = 70
    ws.add_image(logo, f"A{start_row}")
    return start_row + 4


def _write_excel_headers(ws, headers, header_row):
    header_fill = PatternFill(start_color=BRAND_HEADER_COLOR, end_color=BRAND_HEADER_COLOR, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    return header_row + 1


def export_inventory_excel(rows, title="Inventory Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "Company",
        "Ink Type",
        "Opening Stock",
        "Total Received",
        "Total Used",
        "Current Stock",
        "Low Stock Threshold",
        "Status",
    ]

    header_row = _add_excel_logo(ws)
    next_row = _write_excel_headers(ws, headers, header_row)

    for offset, item in enumerate(rows):
        row_num = next_row + offset
        values = [
            item["company"].name,
            item["ink_type"].name,
            item["opening"],
            item["received"],
            item["used"],
            item["current"],
            item["threshold"],
            "LOW" if item["is_low"] else "OK",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_transactions_excel(transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Company", "Ink Type", "Type", "Used", "Left", "Notes", "Entered By", "Created At"]
    header_row = _add_excel_logo(ws)
    next_row = _write_excel_headers(ws, headers, header_row)

    for offset, txn in enumerate(transactions):
        row_num = next_row + offset
        values = [
            txn.transaction_date.strftime("%Y-%m-%d"),
            txn.company.name,
            txn.ink_type.name,
            txn.transaction_type,
            txn.quantity if txn.transaction_type == "Stock Used" else "",
            txn.quantity_left if txn.transaction_type == "Stock Used" else txn.quantity,
            txn.notes or "",
            txn.created_by.username if txn.created_by else "",
            txn.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_inventory_pdf(rows, title="Inventory Report"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH), width=2.2 * inch, height=0.85 * inch, kind="proportional")
        elements.append(logo)
        elements.append(Spacer(1, 0.15 * inch))

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
    )
    elements.append(Paragraph(title, title_style))
    elements.append(
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
    )
    elements.append(Spacer(1, 0.2 * inch))

    data = [
        [
            "Company",
            "Ink Type",
            "Opening",
            "Received",
            "Used",
            "Current",
            "Threshold",
            "Status",
        ]
    ]
    for item in rows:
        data.append(
            [
                item["company"].name,
                item["ink_type"].name,
                f"{item['opening']:.1f}",
                f"{item['received']:.1f}",
                f"{item['used']:.1f}",
                f"{item['current']:.1f}",
                str(item["threshold"]),
                "LOW" if item["is_low"] else "OK",
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_HEADER_COLOR}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
