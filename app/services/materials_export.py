from io import BytesIO
from datetime import datetime

from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from app.services.export import (
    BRAND_HEADER_COLOR,
    LOGO_PATH,
    _add_excel_logo,
    _write_excel_headers,
)
from app.services.materials_inventory import _material_label
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BRAND_RED = colors.HexColor("#B21E22")
BRAND_BLACK = colors.HexColor("#1A1A1A")
ALT_ROW = colors.HexColor("#F9FAFB")
BORDER_GREY = colors.HexColor("#D1D5DB")

PERIOD_TITLES = {
    "daily": "Daily Stock Usage Report",
    "weekly": "Weekly Stock Usage Report",
    "monthly": "Monthly Stock Usage Report",
}


def export_material_inventory_excel(rows, title="Materials Inventory Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    headers = [
        "Material",
        "Opening (kg)",
        "Received (kg)",
        "Used (kg)",
        "Current (kg)",
        "Low Stock Threshold",
        "Status",
    ]

    header_row = _add_excel_logo(ws)
    next_row = _write_excel_headers(ws, headers, header_row)

    for offset, item in enumerate(rows):
        row_num = next_row + offset
        values = [
            item["material"].display_name,
            item["opening"],
            item["received"],
            item["used"],
            item["current"],
            item["threshold"],
            "LOW" if item["is_low"] else "OK",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_material_transactions_excel(transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Date",
        "Material",
        "Type",
        "Used (kg)",
        "Left (kg)",
        "Production / Notes",
        "Entered By",
        "Created At",
    ]
    header_row = _add_excel_logo(ws)
    next_row = _write_excel_headers(ws, headers, header_row)

    for offset, txn in enumerate(transactions):
        row_num = next_row + offset
        values = [
            txn.transaction_date.strftime("%Y-%m-%d"),
            txn.material.display_name,
            txn.transaction_type,
            txn.quantity if txn.transaction_type == "Stock Used" else "",
            txn.quantity_left if txn.transaction_type == "Stock Used" else txn.quantity,
            txn.notes or "",
            txn.created_by.username if txn.created_by else "",
            txn.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_material_inventory_pdf(rows, title="Materials Inventory Report"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH), width=2.2 * inch, height=0.85 * inch, kind="proportional")
        elements.append(logo)
        elements.append(Spacer(1, 0.15 * inch))

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
    )
    elements.append(Paragraph(title, title_style))
    elements.append(
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
    )
    elements.append(Spacer(1, 0.2 * inch))

    data = [
        ["Material", "Opening", "Received", "Used", "Current", "Threshold", "Status"]
    ]
    for item in rows:
        data.append(
            [
                item["material"].display_name,
                f"{item['opening']:.1f}",
                f"{item['received']:.1f}",
                f"{item['used']:.1f}",
                f"{item['current']:.1f}",
                str(item["threshold"]),
                "LOW" if item["is_low"] else "OK",
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_HEADER_COLOR}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_material_usage_pdf(report: dict) -> BytesIO:
    """Generate a branded PDF for materials stock-used records."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.4 * inch,
        bottomMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "UsageTitle",
        parent=styles["Normal"],
        fontSize=16,
        textColor=BRAND_RED,
        fontName="Helvetica-Bold",
        leading=18,
        alignment=TA_LEFT,
    )
    subtitle_style = ParagraphStyle(
        "UsageSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=BRAND_BLACK,
        fontName="Helvetica",
        leading=12,
        alignment=TA_LEFT,
    )
    meta_style = ParagraphStyle(
        "UsageMeta",
        parent=styles["Normal"],
        fontSize=9,
        textColor=BRAND_BLACK,
        fontName="Helvetica-Bold",
        leading=11,
        alignment=TA_RIGHT,
    )
    section_style = ParagraphStyle(
        "UsageSection",
        parent=styles["Normal"],
        fontSize=11,
        textColor=BRAND_RED,
        fontName="Helvetica-Bold",
        leading=13,
        spaceBefore=8,
        spaceAfter=6,
    )
    note_style = ParagraphStyle(
        "UsageNote",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#6B7280"),
        fontName="Helvetica",
        leading=10,
        alignment=TA_CENTER,
    )

    elements = []
    period = report.get("period", "daily")
    title = PERIOD_TITLES.get(period, "Materials Stock Usage Report")
    start_date = report["start_date"]
    end_date = report["end_date"]
    period_label = report["period_label"]

    header_left_rows = [
        [Paragraph(title.upper(), title_style)],
        [Paragraph("RN COLOUR — Printing Materials", subtitle_style)],
        [Paragraph(f"Period: {period_label}", subtitle_style)],
    ]
    if start_date != end_date:
        header_left_rows.append(
            [
                Paragraph(
                    f"From {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}",
                    subtitle_style,
                )
            ]
        )

    header_left = Table(header_left_rows, colWidths=[5.8 * inch])
    header_left.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0)]))

    logo_cell = ""
    if LOGO_PATH.exists():
        logo_cell = Image(str(LOGO_PATH), width=2.1 * inch, height=0.82 * inch, kind="proportional")

    header_right = Table(
        [
            [logo_cell],
            [Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", meta_style)],
        ],
        colWidths=[3.0 * inch],
    )
    header_right.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    header = Table([[header_left, header_right]], colWidths=[6.0 * inch, 3.2 * inch])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(header)
    elements.append(Spacer(1, 0.18 * inch))

    summary_data = [
        ["Total Used (kg)", "Usage Records", "Materials Used", "Average per Record (kg)"],
        [
            f"{report['total_used']:.1f}",
            str(report["record_count"]),
            str(report["material_count"]),
            f"{(report['total_used'] / report['record_count']):.1f}"
            if report["record_count"]
            else "0.0",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[2.3 * inch, 2.3 * inch, 2.3 * inch, 2.3 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_RED),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, BORDER_GREY),
                ("BACKGROUND", (0, 1), (-1, 1), ALT_ROW),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2 * inch))

    if report["by_material"]:
        elements.append(Paragraph("Material-wise Usage Summary", section_style))
        material_data = [["#", "Category", "Material", "Records", "Total Used (kg)"]]
        for index, row in enumerate(report["by_material"], start=1):
            material_data.append(
                [
                    str(index),
                    row["category"],
                    row["material_name"],
                    str(row["record_count"]),
                    f"{row['total_used']:.1f}",
                ]
            )
        material_table = Table(
            material_data,
            colWidths=[0.45 * inch, 0.9 * inch, 4.8 * inch, 0.9 * inch, 1.2 * inch],
            repeatRows=1,
        )
        material_table.setStyle(_usage_table_style(len(material_data)))
        elements.append(material_table)
        elements.append(Spacer(1, 0.18 * inch))

    if period != "daily" and report["by_date"]:
        elements.append(Paragraph("Daily Breakdown", section_style))
        daily_data = [["Date", "Records", "Total Used (kg)"]]
        for row in report["by_date"]:
            daily_data.append(
                [
                    row["date"].strftime("%d-%b-%Y"),
                    str(row["record_count"]),
                    f"{row['total_used']:.1f}",
                ]
            )
        daily_table = Table(
            daily_data,
            colWidths=[2.0 * inch, 1.5 * inch, 1.8 * inch],
            repeatRows=1,
        )
        daily_table.setStyle(_usage_table_style(len(daily_data)))
        elements.append(daily_table)
        elements.append(Spacer(1, 0.18 * inch))

    elements.append(Paragraph("Detailed Usage Records", section_style))
    detail_data = [
        ["Date", "Material", "Used (kg)", "Left (kg)", "Production / Notes", "Entered By"]
    ]
    if report["records"]:
        for txn in report["records"]:
            detail_data.append(
                [
                    txn.transaction_date.strftime("%d-%b-%Y"),
                    _material_label(txn.material, txn.material_id),
                    f"{txn.quantity:.1f}",
                    f"{txn.quantity_left:.1f}" if txn.quantity_left is not None else "—",
                    (txn.notes or "—")[:80],
                    txn.created_by.username if txn.created_by else "—",
                ]
            )
    else:
        detail_data.append(["—", "No usage records in this period", "—", "—", "—", "—"])

    detail_table = Table(
        detail_data,
        colWidths=[1.0 * inch, 2.8 * inch, 0.9 * inch, 0.9 * inch, 2.5 * inch, 1.1 * inch],
        repeatRows=1,
    )
    detail_table.setStyle(_usage_table_style(len(detail_data)))
    elements.append(detail_table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(
        Paragraph(
            "RN COLOUR Stock Management System — Confidential internal usage report",
            note_style,
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _usage_table_style(row_count: int) -> TableStyle:
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, BORDER_GREY),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    if row_count > 1:
        style_commands.append(
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ALT_ROW])
        )
    return TableStyle(style_commands)
